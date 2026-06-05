import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
from datetime import date
import calendar

# ── Config ────────────────────────────────────────────────────────────────────
EXCEL_FILE        = "Demographic-Profile-SUMMARY.xlsm"
SCHOOL_SHEET      = "School Tally"
RESIDENCY_SHEET   = "Residency Tally"

# Daily grid: same for all months — row = day_of_month + 3  (day 1 → row 4)
PROGRAM_COL = {
    # Basic Ed NEW (cols C–H)
    ("New", "G7"): 3,  ("New", "G8"): 4,  ("New", "G9"): 5,
    ("New", "G10"): 6, ("New", "G11"): 7, ("New", "G12"): 8,
    # Basic Ed OLD (cols K–P)
    ("Old", "G7"): 11, ("Old", "G8"): 12, ("Old", "G9"): 13,
    ("Old", "G10"): 14,("Old", "G11"): 15,("Old", "G12"): 16,
    # College NEW (cols S–Y)
    ("New", "BSRT"): 19, ("New", "BSCRIM"): 20, ("New", "BSBA/BSA"): 21,
    ("New", "BSHM"): 22, ("New", "BSTM"): 23, ("New", "EDUC"): 24,
    ("New", "BSIT"): 25,
    # College OLD (cols AB–AH)
    ("Old", "BSRT"): 28, ("Old", "BSCRIM"): 29, ("Old", "BSBA/BSA"): 30,
    ("Old", "BSHM"): 31, ("Old", "BSTM"): 32, ("Old", "EDUC"): 33,
    ("Old", "BSIT"): 34,
}

# ── Per-sheet configuration ────────────────────────────────────────────────────
# Each entry maps a sheet name to its specific cell layout for the
# Other Information section (religion, residency, unspecified).
#
# unspec_type:
#   "single" -> one cumulative cell (May-style)
#               keys: unspec_row, unspec_basic_col, unspec_college_col
#   "daily"  -> one row per day (June-style)
#               keys: unspec_base_row (row = base + day),
#                     unspec_basic_col, unspec_college_col
#
SHEET_CONFIG = {
    "May 2026": {
        # Religion: label col G(7), count col I(9), rows 40-52
        "religion_label_col": 7,
        "religion_count_col": 9,
        "religion_rows": {
            "catholic": 40, "christian": 41, "muslim": 42,
            "iglesia ni cristo": 43, "iglesia": 43,
            "baptist": 44, "babtist": 44,
            "methodist": 45, "adventist": 46, "not specified": 47,
            "jehovah's witness": 48, "jehova": 48,
            "mcgi": 49, "los": 50, "full chapel": 51, "protestant": 52,
        },
        # Residency: label col W(23), count col Z(26), rows 42-65
        "residency_label_col": 23,
        "residency_count_col": 26,
        "residency_rows": {
            "apalit": 42, "balagtas": 43, "baliuag": 44, "bocaue": 45,
            "bulakan": 46, "bustos": 47, "calumpit": 48, "guiguinto": 49,
            "hagonoy": 50, "malolos": 51, "manila": 52, "marilao": 53,
            "meycauayan": 54, "santa maria": 55, "pampanga": 56, "pandi": 57,
            "paombong": 58, "plaridel": 59, "pulilan": 60,
            "san ildefonso": 61, "san miguel": 62, "san rafael": 63,
            "drt": 64, "others": 65,
        },
        # Unspecified: single cumulative row
        "unspec_type": "single",
        "unspec_row": 38,
        "unspec_basic_col": 29,   # col AC
        "unspec_college_col": 32, # col AF
    },

    "June 2026": {
        # Religion: label col C(3), count col E(5), rows 40-53
        # (count col is empty in the sheet — app writes here)
        "religion_label_col": 3,
        "religion_count_col": 7,  # col G — outside merged label range C:F
        "religion_rows": {
            "catholic": 40, "christian": 41, "muslim": 42,
            "iglesia ni cristo": 43, "iglesia": 43,
            "baptist": 44, "babtist": 44,
            "methodist": 45, "adventist": 46,
            "jehovah's witness": 47, "jehova": 47,
            "mcgi": 48, "los": 49, "full chapel": 50, "protestant": 51,
            "alipay": 52, "not specified": 53,
        },
        # Residency: label col AB(28), count col AD(30), rows 40-63
        "residency_label_col": 28,
        "residency_count_col": 30,
        "residency_rows": {
            "apalit": 40, "balagtas": 41, "baliuag": 42, "bocaue": 43,
            "bulakan": 44, "bustos": 45, "calumpit": 46, "guiguinto": 47,
            "hagonoy": 48, "malolos": 49, "manila": 50, "marilao": 51,
            "meycauayan": 52, "santa maria": 53, "pampanga": 54, "pandi": 55,
            "paombong": 56, "plaridel": 57, "pulilan": 58,
            "san ildefonso": 59, "san miguel": 60, "san rafael": 61,
            "drt": 62, "others": 63,
        },
        # Unspecified: per-day rows (row = base + day_of_month)
        # confirmed formulas: =SUM(AN40:AN69) and =SUM(AQ40:AQ69)
        "unspec_type": "daily",
        "unspec_base_row": 39,    # row 40 = day 1 → base(39) + 1
        "unspec_basic_col": 40,   # col AN
        "unspec_college_col": 43, # col AQ
    },
}

PROGRAMS   = ["-- Select --", "BSRT", "BSCRIM", "BSBA/BSA", "BSHM",
              "BSTM", "EDUC", "BSIT", "G7", "G8", "G9", "G10", "G11", "G12",
              "Unspecified"]
RESIDENCES = ["Pulilan", "Plaridel", "Malolos", "Baliuag", "Guiguinto", "Others"]
RELIGIONS  = ["Catholic", "Christian", "Muslim", "Iglesia ni Cristo", "Baptist",
              "Methodist", "Adventist", "Jehovah's Witness", "MCGI", "LOS",
              "Full Chapel", "Protestant", "Not specified", "Others"]

# ── Excel helpers ─────────────────────────────────────────────────────────────
def get_sheet_name():
    t = date.today()
    return f"{calendar.month_name[t.month]} {t.year}"

def get_config(sheet_name):
    """Return sheet config, or None if sheet is unknown."""
    return SHEET_CONFIG.get(sheet_name)

def increment_cell(ws, row, col):
    current = ws.cell(row, col).value
    ws.cell(row, col).value = (current or 0) + 1

def get_all_schools(wb):
    """Read all existing school names for autocomplete."""
    schools = set()
    sheet = get_sheet_name()
    if sheet in wb.sheetnames:
        ws = wb[sheet]
        cfg = get_config(sheet)
        # Previous school list columns differ per sheet
        school_cols = [12, 19] if cfg and cfg.get("religion_label_col") == 3 else [11, 17]
        for r in range(40, 200):
            for col in school_cols:
                v = ws.cell(r, col).value
                if v and isinstance(v, str) and v.strip():
                    schools.add(v.strip())
    if SCHOOL_SHEET in wb.sheetnames:
        ws = wb[SCHOOL_SHEET]
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v and isinstance(v, str) and v.strip():
                schools.add(v.strip())
    return sorted(schools)

def setup_school_sheet(wb):
    """Create School Tally sheet if it doesn't exist."""
    if SCHOOL_SHEET in wb.sheetnames:
        return
    ws = wb.create_sheet(SCHOOL_SHEET)
    for col, text in [(1, "Previous School"), (2, "Count")]:
        cell = ws.cell(1, col, text)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 10
    ws.freeze_panes = "A2"

def setup_residency_sheet(wb):
    """Create Residency Tally sheet if it doesn't exist."""
    if RESIDENCY_SHEET in wb.sheetnames:
        return
    ws = wb.create_sheet(RESIDENCY_SHEET)
    for col, text in [(1, "Municipality / Residency"), (2, "Count")]:
        cell = ws.cell(1, col, text)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10
    ws.freeze_panes = "A2"

def get_all_residencies(wb):
    """Read all existing municipality names for autocomplete."""
    names = set()
    if RESIDENCY_SHEET in wb.sheetnames:
        ws = wb[RESIDENCY_SHEET]
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v and isinstance(v, str) and v.strip():
                names.add(v.strip())
    # Seed with known municipalities if sheet is empty
    if not names:
        names = {
            "Pulilan", "Plaridel", "Malolos", "Baliuag", "Guiguinto",
            "Apalit", "Balagtas", "Bocaue", "Bulakan", "Bustos",
            "Calumpit", "Hagonoy", "Manila", "Marilao", "Meycauayan",
            "Santa Maria", "Pampanga", "Pandi", "Paombong",
            "San Ildefonso", "San Miguel", "San Rafael", "DRT",
        }
    return sorted(names)

def tally_residency(wb, municipality):
    """Increment municipality count in Residency Tally sheet."""
    if not municipality.strip():
        return
    ws = wb[RESIDENCY_SHEET]
    name_clean = municipality.strip()
    for r in range(2, ws.max_row + 1):
        existing = ws.cell(r, 1).value
        if existing and existing.strip().lower() == name_clean.lower():
            ws.cell(r, 2).value = (ws.cell(r, 2).value or 0) + 1
            return
    new_row = ws.max_row + 1
    ws.cell(new_row, 1).value = name_clean
    ws.cell(new_row, 1).font  = Font(name="Arial", size=10)
    ws.cell(new_row, 2).value = 1
    ws.cell(new_row, 2).font  = Font(name="Arial", size=10)
    ws.cell(new_row, 2).alignment = Alignment(horizontal="center")

def tally_school(wb, school_name):
    """Increment school count in School Tally sheet, adding new row if needed."""
    if not school_name.strip():
        return
    ws = wb[SCHOOL_SHEET]
    name_clean = school_name.strip()
    for r in range(2, ws.max_row + 1):
        existing = ws.cell(r, 1).value
        if existing and existing.strip().lower() == name_clean.lower():
            ws.cell(r, 2).value = (ws.cell(r, 2).value or 0) + 1
            return
    new_row = ws.max_row + 1
    ws.cell(new_row, 1).value = name_clean
    ws.cell(new_row, 1).font  = Font(name="Arial", size=10)
    ws.cell(new_row, 2).value = 1
    ws.cell(new_row, 2).font  = Font(name="Arial", size=10)
    ws.cell(new_row, 2).alignment = Alignment(horizontal="center")

def tally(student_type, program, religion, residence, school=None, unspec_level=None):
    wb   = load_workbook(EXCEL_FILE, keep_vba=True)
    sheet = get_sheet_name()

    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{sheet}' not found.\n"
            f"Available sheets: {', '.join(wb.sheetnames)}"
        )

    cfg = get_config(sheet)
    if cfg is None:
        wb.close()
        raise ValueError(
            f"Sheet '{sheet}' is not configured in this app yet.\n"
            f"Configured sheets: {', '.join(SHEET_CONFIG.keys())}\n\n"
            f"Please contact your administrator to add support for this month."
        )

    ws        = wb[sheet]
    today     = date.today()
    today_row = today.day + 3   # row 4 = day 1

    # 1 ── Daily program tally
    if program == "Unspecified" and unspec_level:
        if cfg["unspec_type"] == "single":
            col = cfg["unspec_basic_col"] if unspec_level == "Basic Ed" else cfg["unspec_college_col"]
            increment_cell(ws, cfg["unspec_row"], col)
        else:  # daily
            row = cfg["unspec_base_row"] + today.day
            col = cfg["unspec_basic_col"] if unspec_level == "Basic Ed" else cfg["unspec_college_col"]
            increment_cell(ws, row, col)
    else:
        key = (student_type, program)
        if key in PROGRAM_COL:
            increment_cell(ws, today_row, PROGRAM_COL[key])

    # 2 ── Religion tally
    if religion and religion.lower() not in ("others", ""):
        rel_key = religion.lower().strip()
        matched_row = None
        for k, r in cfg["religion_rows"].items():
            if k in rel_key or rel_key in k:
                matched_row = r
                break
        if matched_row:
            increment_cell(ws, matched_row, cfg["religion_count_col"])
        else:
            # New religion not in list — append after last known religion row
            label_col = cfg["religion_label_col"]
            count_col = cfg["religion_count_col"]
            max_known = max(cfg["religion_rows"].values()) + 1
            for r in range(max_known, max_known + 20):
                if ws.cell(r, label_col).value is None:
                    ws.cell(r, label_col).value = religion
                    ws.cell(r, count_col).value = 1
                    break

    # 3 ── Residency tally
    if residence:
        res_key = residence.lower().replace("others:", "others").strip()
        if res_key.startswith("others"):
            res_key = "others"
        res_row = cfg["residency_rows"].get(res_key)
        if res_row:
            increment_cell(ws, res_row, cfg["residency_count_col"])

    # 4 ── Previous school tally (separate sheet — never touches summary layout)
    if school and school.strip():
        setup_school_sheet(wb)
        tally_school(wb, school)

    # 5 ── Residency tally (separate sheet)
    if residence and residence.strip():
        setup_residency_sheet(wb)
        tally_residency(wb, residence)

    wb.save(EXCEL_FILE)
    wb.close()


# ── GUI ───────────────────────────────────────────────────────────────────────
class TallyApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Enrollment Tally — OLM Pulilan")
        self.root.resizable(False, False)

        if not os.path.exists(EXCEL_FILE):
            messagebox.showerror("File Not Found",
                f"'{EXCEL_FILE}' not found.\n"
                "Place this script in the same folder as the Excel file.")
            root.destroy()
            return

        sheet = get_sheet_name()
        wb = load_workbook(EXCEL_FILE, keep_vba=True)

        # Warn if sheet or config is missing
        if sheet not in wb.sheetnames:
            messagebox.showwarning("Sheet Missing",
                f"No sheet named '{sheet}' found.\n"
                f"Available: {', '.join(wb.sheetnames)}\n\n"
                "Saving will fail until the correct sheet exists.")
            self.school_list    = []
            self.residency_list = []
        elif sheet not in SHEET_CONFIG:
            messagebox.showwarning("Sheet Not Configured",
                f"Sheet '{sheet}' exists but has no configuration in this app.\n\n"
                "Daily enrollment tally will still work.\n"
                "Religion, Residency, and Unspecified tallying will be skipped.")
            self.school_list    = get_all_schools(wb)
            self.residency_list = get_all_residencies(wb)
        else:
            self.school_list     = get_all_schools(wb)
            self.residency_list  = get_all_residencies(wb)
        wb.close()

        self._build_styles()
        self._build_ui()

    # ── Styles ────────────────────────────────────────────────────────────────
    def _build_styles(self):
        s = ttk.Style(self.root)
        s.theme_use("clam")
        s.configure("Title.TLabel",   font=("Arial", 9),          foreground="#555")
        s.configure("Head.TLabel",    font=("Arial", 12, "bold"),  foreground="#1F4E79")
        s.configure("Section.TLabel", font=("Arial", 9, "bold"),   foreground="#1F4E79")
        s.configure("Sheet.TLabel",   font=("Arial", 9),           foreground="#888")
        s.configure("TLabel",         font=("Arial", 10))
        s.configure("TEntry",         font=("Arial", 10))
        s.configure("TCombobox",      font=("Arial", 10))
        s.configure("Save.TButton",   font=("Arial", 11, "bold"),
                    background="#1F4E79", foreground="white", padding=10)
        s.configure("Clear.TButton",  font=("Arial", 10),
                    background="#E0E0E0", padding=10)
        s.map("Save.TButton",  background=[("active", "#2E74B5")])
        s.map("Clear.TButton", background=[("active", "#CCCCCC")])

    # ── UI Layout ─────────────────────────────────────────────────────────────
    def _build_ui(self):
        pad = ttk.Frame(self.root, padding=24)
        pad.pack(fill="both", expand=True)

        # Header
        ttk.Label(pad, text="College of Our Lady of Mercy of Pulilan Foundation Inc.",
                  style="Title.TLabel").grid(row=0, column=0, columnspan=4)
        ttk.Label(pad, text="ENROLLMENT TALLY",
                  style="Head.TLabel").grid(row=1, column=0, columnspan=4, pady=(2, 4))

        today  = date.today()
        sheet  = get_sheet_name()
        cfg_ok = sheet in SHEET_CONFIG
        status_color = "#888" if cfg_ok else "#c0392b"
        status_text  = (f"{today.strftime('%A, %B %d, %Y')}  ·  Sheet: {sheet}"
                        if cfg_ok else
                        f"{today.strftime('%A, %B %d, %Y')}  ·  Sheet: {sheet}  ⚠ not configured")
        ttk.Label(pad, text=status_text,
                  font=("Arial", 9), foreground=status_color).grid(
            row=2, column=0, columnspan=4, pady=(0, 12))
        ttk.Separator(pad).grid(row=3, column=0, columnspan=4, sticky="ew", pady=(0, 14))

        self.v = {}
        r = 4

        # ── Enrollment ────────────────────────────────────────────────────
        ttk.Label(pad, text="ENROLLMENT", style="Section.TLabel").grid(
            row=r, column=0, columnspan=4, sticky="w", pady=(0, 8))
        r += 1

        self._lbl(pad, "Student Type:", r, 0)
        sf = ttk.Frame(pad); sf.grid(row=r, column=1, sticky="w", padx=(6, 20))
        self.v["student_type"] = tk.StringVar()
        for val in ("New", "Old"):
            ttk.Radiobutton(sf, text=f"{val} Student",
                            variable=self.v["student_type"], value=val).pack(side="left", padx=(0,10))

        self._lbl(pad, "Intended Program:", r, 2)
        self.v["program"] = tk.StringVar(value="-- Select --")
        self.prog_box = ttk.Combobox(pad, textvariable=self.v["program"],
                                     values=PROGRAMS, width=18, state="readonly")
        self.prog_box.grid(row=r, column=3, sticky="w", padx=(6, 0))
        self.prog_box.bind("<<ComboboxSelected>>", self._on_program_change)
        r += 1

        # Unspecified level (hidden until program=Unspecified)
        self._lbl(pad, "Level (Unspecified):", r, 0)
        self.v["unspec_level"] = tk.StringVar()
        self.unspec_frame = ttk.Frame(pad)
        self.unspec_frame.grid(row=r, column=1, sticky="w", padx=(6, 0))
        for val in ("Basic Ed", "College"):
            ttk.Radiobutton(self.unspec_frame, text=val,
                            variable=self.v["unspec_level"], value=val).pack(side="left", padx=(0, 10))
        self.unspec_row_idx = r
        r += 1
        self._toggle_unspec(False)

        ttk.Separator(pad).grid(row=r, column=0, columnspan=4, sticky="ew", pady=10)
        r += 1

        # ── Demographics ──────────────────────────────────────────────────
        ttk.Label(pad, text="DEMOGRAPHICS  (religion & residence)",
                  style="Section.TLabel").grid(row=r, column=0, columnspan=4, sticky="w", pady=(0,8))
        r += 1

        self._lbl(pad, "Religion:", r, 0)
        self.v["religion"] = tk.StringVar()
        ttk.Combobox(pad, textvariable=self.v["religion"],
                     values=RELIGIONS, width=22).grid(
            row=r, column=1, sticky="w", padx=(6, 20))

        self._lbl(pad, "Residence:", r, 2)
        self.v["residence"] = tk.StringVar()
        self.res_entry = ttk.Combobox(pad, textvariable=self.v["residence"],
                                      values=self.residency_list, width=22)
        self.res_entry.grid(row=r, column=3, sticky="w", padx=(6, 0))
        self.res_entry.bind("<KeyRelease>", self._filter_residencies)
        r += 1

        ttk.Label(pad, text="Type to search or enter a new municipality",
                  font=("Arial", 8), foreground="#999").grid(
            row=r, column=3, sticky="w", padx=(6, 0))
        r += 1

        # Previous School
        self._lbl(pad, "Previous School:", r, 0)
        self.v["school"] = tk.StringVar()
        self.school_entry = ttk.Combobox(pad, textvariable=self.v["school"],
                                          values=self.school_list, width=38)
        self.school_entry.grid(row=r, column=1, columnspan=3, sticky="w", padx=(6, 0))
        self.school_entry.bind("<KeyRelease>", self._filter_schools)
        ttk.Label(pad, text="Type to search or enter a new school name",
                  font=("Arial", 8), foreground="#999").grid(
            row=r+1, column=1, columnspan=3, sticky="w", padx=(6, 0))
        r += 2

        ttk.Separator(pad).grid(row=r, column=0, columnspan=4, sticky="ew", pady=14)
        r += 1

        # Buttons
        bf = ttk.Frame(pad); bf.grid(row=r, column=0, columnspan=4)
        ttk.Button(bf, text="✚   Add to Tally", style="Save.TButton",
                   command=self._save).pack(side="left", padx=8)
        ttk.Button(bf, text="🗑   Clear", style="Clear.TButton",
                   command=self._clear).pack(side="left", padx=8)
        r += 1

        self.count_var = tk.StringVar(value="Entries today: 0")
        ttk.Label(pad, textvariable=self.count_var,
                  font=("Arial", 11, "bold"), foreground="#1F4E79").grid(
            row=r, column=0, columnspan=4, pady=(14, 0))
        r += 1

        self.status = tk.StringVar(value="Ready — select program, religion, and residence then click Add to Tally.")
        ttk.Label(pad, textvariable=self.status, foreground="gray",
                  font=("Arial", 9, "italic"), wraplength=500).grid(
            row=r, column=0, columnspan=4, pady=(4, 0))

        self.session_count = 0

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _filter_residencies(self, _=None):
        typed = self.v["residence"].get().lower()
        if typed:
            filtered = [r for r in self.residency_list if typed in r.lower()]
            self.res_entry["values"] = filtered if filtered else self.residency_list
        else:
            self.res_entry["values"] = self.residency_list

    def _filter_schools(self, _=None):
        typed = self.v["school"].get().lower()
        if typed:
            filtered = [s for s in self.school_list if typed in s.lower()]
            self.school_entry["values"] = filtered if filtered else self.school_list
        else:
            self.school_entry["values"] = self.school_list

    def _lbl(self, parent, text, row, col):
        ttk.Label(parent, text=text).grid(row=row, column=col, sticky="w",
                                          pady=5, padx=(0 if col == 0 else 6, 0))

    def _on_program_change(self, _=None):
        self._toggle_unspec(self.v["program"].get() == "Unspecified")

    def _toggle_unspec(self, show: bool):
        if show:
            self.unspec_frame.grid()
        else:
            self.unspec_frame.grid_remove()
            self.v["unspec_level"].set("")

    # ── Actions ───────────────────────────────────────────────────────────────
    def _save(self):
        student_type = self.v["student_type"].get()
        program      = self.v["program"].get()
        religion     = self.v["religion"].get().strip()
        residence    = self.v["residence"].get().strip()
        unspec_level = self.v["unspec_level"].get()

        missing = []
        if not student_type:
            missing.append("Student Type")
        if program in ("", "-- Select --"):
            missing.append("Intended Program")
        if program == "Unspecified" and not unspec_level:
            missing.append("Level for Unspecified (Basic Ed or College)")
        if missing:
            messagebox.showwarning("Required Fields",
                                   "Please fill in:\n• " + "\n• ".join(missing))
            return

        try:
            tally(student_type, program, religion, residence,
                  school=self.v["school"].get().strip(),
                  unspec_level=unspec_level if program == "Unspecified" else None)

            self.session_count += 1
            self.count_var.set(f"Entries today: {self.session_count}")
            prog_label = f"{program} ({unspec_level})" if program == "Unspecified" else program
            self.status.set(
                f"✅  Tallied: {student_type} Student · {prog_label}"
                + (f"  · {religion}" if religion else "")
                + (f"  · {residence}" if residence else "")
                + (f"  · {self.v['school'].get().strip()}" if self.v["school"].get().strip() else "")
            )
            wb = load_workbook(EXCEL_FILE, keep_vba=True)
            self.school_list    = get_all_schools(wb)
            self.residency_list = get_all_residencies(wb)
            wb.close()
            self.school_entry["values"] = self.school_list
            self.res_entry["values"]    = self.residency_list
            self._clear()
        except Exception as e:
            messagebox.showerror("Save Error",
                f"Could not update Excel:\n{e}\n\n"
                "Make sure the file is not open in Excel.")

    def _clear(self):
        self.v["student_type"].set("")
        self.v["program"].set("-- Select --")
        self.v["religion"].set("")
        self.v["residence"].set("")
        self.v["school"].set("")
        self.v["unspec_level"].set("")
        self._toggle_unspec(False)


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    TallyApp(root)
    root.mainloop()
